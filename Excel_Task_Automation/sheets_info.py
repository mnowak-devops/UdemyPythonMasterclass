import openpyxl
#Loading files
file_path = r"C:\Udemy_Python_Masterclass\Excel_Task_Automation\Employees.xlsx"
workbook = openpyxl.load_workbook(file_path)

sheet = workbook['EmployeeData']

print(sheet.title)
print("\n")

print(dir (sheet)) #Available comments on sheet
print("\n")

print(sheet.max_row)
print("\n")


print(sheet.max_column)
print("\n")


for i in sheet.values:
    print (i)


"""
#Sheets
#Getting the title of a sheet
sheet.title

#Setting the title of a sheet
sheet.title = 'Employees'

#Renaming the default sheet to 'Employees'
sheet = workbook['Sheet']

sheet.title = 'Employees'

#Seeing all the available sheet operations
dir(sheet)

#Getting the active cell
sheet.active_cell

#Getting the dimensions of the sheet (the array populated with data)
sheet.dimensions

#Getting basic parameters about the sheet
sheet.sheet_format

#Getting the basic properties of the sheet
sheet.sheet_properties

#Getting the sheet state (visible/hidden)
sheet.sheet_state

#Getting other general information about a sheet
sheet.sheet_view

#Getting the number of rows in the sheet
sheet.max_row

#Getting the number of columns in the sheet
sheet.max_column

#Returning the rows in a sheet as tuples
for i in sheet.values:
	print(i)

 """