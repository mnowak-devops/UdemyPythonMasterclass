import openpyxl

file_path = r"C:\Udemy_Python_Masterclass\Excel_Task_Automation\Employees.xlsx"
workbook = openpyxl.load_workbook(r"C:\Udemy_Python_Masterclass\Excel_Task_Automation\Employees.xlsx")

print (workbook.properties)

print (workbook.sheetnames)

print (workbook.active)

sheet = workbook['EmployeeData']

workbook.create_sheet('TestSheet')

workbook.save(file_path)

print (workbook.sheetnames)

sheet = workbook['TestSheet']

workbook.remove(sheet)


# del workbook['TestSheet']
# Save a file
workbook.save(r"C:\Udemy_Python_Masterclass\Excel_Task_Automation\Employees.xlsx")

print (workbook.sheetnames)

#Closing a workbook
workbook.close()


"""
#Workbooks
#Creating a new workbook
from openpyxl import Workbook

workbook = Workbook() #Creating a new workbook object, by initializing the Workbook() class

file_path = r"D:\\ExcelApp\\MyCompanyStaff.xlsx" #Setting the path to (location of) the new Excel workbook

workbook.save(file_path) #Saving the workbook

#Loading the Excel workbook .xlsx file
workbook = openpyxl.load_workbook(r"D:\Employees.xlsx")

#Getting the basic properties of the workbook
workbook.properties

#Getting the sheets in the workbook
workbook.worksheets

#Getting the sheet names in a workbook, as a list of strings
workbook.sheetnames

#Getting the active sheet in the workbook
workbook.active

#Referencing a sheet by its name
workbook['EmployeeData']

sheet = workbook['EmployeeData']

#Creating a new sheet in the workbook
workbook.create_sheet('TestSheet')
workbook.save(r"D:\Employees.xlsx") #Saving the workbook

#Removing a sheet from the workbook
sheet = workbook['TestSheet']
workbook.remove(sheet)

#or

del workbook['TestSheet']

workbook.save(r"D:\Employees.xlsx") #Saving the workbook

#Closing a workbook
workbook.close()

"""