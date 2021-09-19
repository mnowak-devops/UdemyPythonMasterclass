import openpyxl
#Loading files
file_path = r"C:\Udemy_Python_Masterclass\Excel_Task_Automation\Employees.xlsx"
workbook = openpyxl.load_workbook(file_path)

sheet = workbook['EmployeeData']


print(sheet['B7'].value)

# Wrong comand sheet
#sheet.cell(row = 6, column = 2),value


print(sheet.cell(row = 6, column = 2).value)

call = sheet['B9']


print(call.row)
print(call.column)
print(call.coordinate)
print(call.data_type)

print(call.encoding)

call = sheet['B2']
call.value = 'David'
workbook.save(file_path)

print(call.parent)


"""
#Getting the value of a cell in a sheet, by the cell coordinates
sheet = workbook['EmployeeData']
sheet['B10'].value

#or

sheet.cell(row = 10, column = 2).value

#Selecting a cell in the sheet
cell = sheet['B10']

#Getting the row number of the cell
cell.row

#Getting the column number of the cell
cell.column

#Getting the coordinates of a cell
cell.coordinate

#Getting the data type for a cell
'''
TYPE_STRING = 's'
    TYPE_FORMULA = 'f'
    TYPE_NUMERIC = 'n'
    TYPE_BOOL = 'b'
    TYPE_NULL = 'n'
    TYPE_INLINE = 'inlineStr'
    TYPE_ERROR = 'e'
    TYPE_FORMULA_CACHE_STRING = 'str'
'''
cell.data_type

#Getting the encoding format for a cell
cell.encoding

#Setting a value for a cell
cell = sheet['B10']
cell.value = 'Test'
workbook.save("D:\Employees.xlsx") #Saving the workbook

#Getting the parent worksheet for a cell
cell.parent

"""